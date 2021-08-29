import openpyxl
import json,time
from openpyxl import load_workbook

class ReadExecl:
    def __init__(self,filename,sheet):
        self.filepath='../data/'+filename
        self.wb=load_workbook(self.filepath)
        self.sheet = self.wb[sheet]

    # 测试一条用例
    # def read_execl(self):
    #     row_data = {}  # 第一行的数据为空数据
    #     '''把第一行每个单元格的数据以key:value的形式存到row_data这个知道里面'''
    #     row_data['case_id']=self.sheet.cell(2,1).value#行由i确定
    #     row_data['title']=self.sheet.cell(2,2).value
    #     row_data['url']=self.sheet.cell(2, 3).value
    #     row_data['param']=self.sheet.cell(2, 4).value
    #     row_data['method']=self.sheet.cell(2, 5).value
    #     row_data['expect_code']=self.sheet.cell(2, 6).value
    #     row_data['message'] = self.sheet.cell(2, 7).value
    #     # print(row_data['param'],type(row_data['param']))
    #     # print(eval(row_data['param']))
    #     return [eval(row_data['param']),row_data['message']]



    # 测试多条用例
    def read_execl(self):
        test_data=[]
        test_data_finlly=[]
        for i in range(2, self.sheet.max_row + 1):  # 按行获取
            row_data={}  #每一行的数据为空数据
            '''把每一行每个单元格的数据以key:value的形式存到row_data这个知道里面'''
            row_data['case_id']=self.sheet.cell(i,1).value#行由i确定
            row_data['title']=self.sheet.cell(i,2).value
            row_data['url']=self.sheet.cell(i, 3).value
            row_data['param']=self.sheet.cell(i, 4).value
            row_data['method']=self.sheet.cell(i, 5).value
            row_data['expect_code']=self.sheet.cell(i, 6).value
            row_data['message'] = self.sheet.cell(i, 7).value
            test_data.append(row_data)
        # print(test_data)
        for one in test_data:
            test_data_finlly.append([eval(one['param']),one['message']])
        return test_data_finlly




if __name__ == '__main__':
    datas=ReadExecl('cases.xlsx','add_message').read_execl()
    print(datas)
    # print(str(datas),type(str(datas)))
